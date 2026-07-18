"""Generate Accuracy_Improvement_Plan.xlsx (combined roadmap)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "Accuracy_Improvement_Plan.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUB_FONT = Font(bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PRIORITY_FILL = {
    "P0": PatternFill("solid", fgColor="FCE4D6"),
    "P1": PatternFill("solid", fgColor="FFF2CC"),
    "P2": PatternFill("solid", fgColor="E2EFDA"),
    "P3": PatternFill("solid", fgColor="F2F2F2"),
    "Done": PatternFill("solid", fgColor="C6E0B4"),
}


def style_header_row(ws, row: int, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def write_table(ws, headers, rows, start_row=1, priority_col=None):
    ncol = len(headers)
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header_row(ws, start_row, ncol)
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
            if priority_col and c_idx == priority_col:
                key = str(val).split()[0] if val else ""
                if key in PRIORITY_FILL:
                    cell.fill = PRIORITY_FILL[key]
    return start_row + len(rows) + 2


def autosize(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 0
        for cell in col:
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = max(10, width)


def build_summary(ws):
    ws.title = "Summary"
    ws["A1"] = "Change Detection — Accuracy Improvement Roadmap"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    lines = [
        ("Recommended approach", "Hybrid: keep completed classical/infrastructure work; "
         "make Delhi-labeled evaluation the single source of truth; fine-tune AdaptFormer "
         "for domain transfer; calibrate fusion/thresholds against real IoU. "
         "Defer extended KPCAMNet and new model training until fine-tuning results are measured."),
        ("Why this wins", "Prior integration (KPCA, IR-MAD, GSD, benchmark harness) improved "
         "machinery and regression safety, but LEVIR-CD tiles do not represent Delhi imagery. "
         "Domain mismatch remains the dominant accuracy gap. Labeled Delhi pairs + transfer "
         "learning addresses the root cause; grid-search calibration is fast, measurable, and "
         "complements both."),
        ("Non-negotiable first step", "Build a Delhi evaluation set (20–40 pairs minimum to "
         "start; grow to 100–300 for fine-tuning). Without it, no improvement can be verified."),
        ("Highest expected payoff", "Fine-tune AdaptFormer on Delhi imagery (GPU, 1–2 weeks "
         "engineering + labeling time)."),
        ("Explicitly deprioritize", "Siamese U-Net from scratch (needs 1000+ labeled pairs). "
         "Extended KPCAMNet only if unsupervised fallback is required."),
        ("Target success metrics", "Delhi eval mean F1 ≥ 0.55 (phase 1 baseline target); "
         "≥ 0.65 after calibration; ≥ 0.70 after fine-tuning. Car-FP synthetic gate stays "
         "F1 = 1.0 (no regression)."),
        ("Total timeline (estimate)", "4–6 weeks end-to-end: 1 week eval + calibration, "
         "2–3 weeks labeling expansion + fine-tune, 1 week integration + validation."),
    ]
    row = 3
    for title, body in lines:
        ws.cell(row=row, column=1, value=title).font = SUB_FONT
        ws.cell(row=row, column=2, value=body).alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    for r in range(3, row):
        ws.row_dimensions[r].height = 42


def build_comparison(ws):
    ws.title = "Approach Comparison"
    headers = [
        "Topic", "Prior Plan (SRCDNet/KPCAMNet/CD-Review)",
        "Found Plan (Delhi-focused)", "Assessment", "Combined Decision",
    ]
    rows = [
        ("Evaluation data", "LEVIR-CD sample tiles + synthetic cases",
         "20–40 Delhi pairs with hand-drawn masks",
         "LEVIR useful for regression; useless as Delhi proxy",
         "Keep LEVIR/synthetic as CI gates; Delhi set = primary metric"),
        ("Domain adaptation", "Not addressed (pretrained AdaptFormer only)",
         "Fine-tune AdaptFormer on Delhi (transfer learning)",
         "Biggest gap for your imagery",
         "P0: build fine-tuning pipeline + Delhi weights"),
        ("Threshold / fusion tuning", "Manual env flags + validation gates",
         "Grid-search sensitivity/fusion/KPCA params on real IoU",
         "Fast, no training, immediate gain",
         "P0: add compare_methods.py + calibration script"),
        ("KPCAMNet / KPCA", "Implemented basic KPCA + polar analysis",
         "Extend multi-channel KPCA (lower priority)",
         "Helped Feature-Based (F1 0.12→0.32 on LEVIR); not AI-path fix",
         "Keep current KPCA; extend only if unsupervised path needed (P2)"),
        ("IR-MAD", "Implemented channel + optional regression norm",
         "Not emphasized",
         "Regression norm hurt F1; channel neutral on Delhi gate",
         "Keep channel opt-in; skip default regression norm"),
        ("GSD harmonization", "Implemented for mismatched GeoTIFF GSD",
         "Not mentioned",
         "Prevents false texture diffs on mixed-resolution pairs",
         "Keep enabled (already done)"),
        ("BIT_CD ensemble", "Optional second DL model",
         "Not mentioned",
         "Needs weights; payoff unproven on Delhi",
         "Defer until after AdaptFormer fine-tune (P3)"),
        ("Siamese U-Net training", "Explicitly skipped",
         "Not recommended near-term",
         "Agree — data-hungry",
         "Skip unless 1000+ labeled pairs available"),
        ("Car / transient FP guard", "Synthetic parked_cars benchmark case",
         "Not mentioned",
         "Critical for Delhi road imagery",
         "Keep as mandatory regression gate"),
    ]
    end = write_table(ws, headers, rows, start_row=1, priority_col=None)
    ws.cell(row=end, column=1, value="Verdict").font = SUB_FONT
    ws.merge_cells(start_row=end, start_column=1, end_row=end, end_column=5)
    ws.cell(row=end, column=1,
            value="Verdict: Found plan correctly identifies the bottleneck (domain + measurement). "
                  "Prior plan correctly built reusable infrastructure. Combined roadmap = Delhi "
                  "eval first → calibrate → fine-tune → integrate.").alignment = WRAP
    autosize(ws)


def build_plan(ws):
    ws.title = "Implementation Plan"
    headers = [
        "Phase", "Priority", "Workstream", "Task", "Description",
        "Depends On", "Effort", "Owner", "Deliverable", "Success Criteria", "Status",
    ]
    rows = [
        ("0", "Done", "Foundation", "Benchmark harness",
         "validate_detection.py: synthetic + LEVIR-CD + kappa + car-FP gate",
         "—", "Done", "Eng", "scripts/validate_detection.py",
         "All unit checks pass; synthetic F1 stable", "Complete"),
        ("0", "Done", "Foundation", "Classical channels",
         "KPCA (feature path), IR-MAD (opt-in), GSD harmonization, BIT_CD scaffold",
         "—", "Done", "Eng", "app/cd_models/*, detection_engine.py",
         "No car-FP regression; infra ready", "Complete"),
        ("1", "P0", "Evaluation", "Curate Delhi image pairs",
         "Select 20–40 representative before/after pairs from library (building, road, "
         "open land, mixed GSD). Document pair metadata (date, GSD, zone).",
         "—", "2–4 days", "Domain + Eng", "docs/delhi_eval/manifest.json",
         "≥20 pairs covering main change types", "Not started"),
        ("1", "P0", "Evaluation", "Create ground-truth masks",
         "Hand-label binary change masks in QGIS or LabelMe (rough polygons OK). "
         "One mask per pair, same resolution as detection input.",
         "Curate Delhi pairs", "2–4 days", "Domain", "docs/delhi_eval/labels/*.png",
         "Every pair has aligned GT mask", "Not started"),
        ("1", "P0", "Evaluation", "Wire Delhi eval into harness",
         "Extend validate_detection.py --real to load docs/delhi_eval/; report per-pair "
         "and mean IoU/F1/kappa; save comparison PNGs.",
         "GT masks", "0.5 day", "Eng", "Updated validate_detection.py",
         "Baseline Delhi metrics recorded", "Not started"),
        ("1", "P0", "Evaluation", "Record baseline report",
         "Run AI-Based DL + Feature-Based + Hybrid at default settings; store metrics.json "
         "as baseline for all future A/B tests.",
         "Delhi eval wired", "0.5 day", "Eng", "runs/delhi_baseline/",
         "Baseline F1 documented per method", "Not started"),
        ("2", "P0", "Calibration", "Build compare_methods.py",
         "New script: sweep methods, fusion modes, sensitivity, DETECTION_* flags; "
         "output ranked CSV/JSON vs Delhi GT.",
         "Delhi eval", "1 day", "Eng", "scripts/compare_methods.py",
         "Single command produces method leaderboard", "Not started"),
        ("2", "P0", "Calibration", "Grid-search fusion thresholds",
         "Search smart_union floors, hysteresis high/low, classical percentile q, "
         "DL threshold, KPCA on/off for score map.",
         "compare_methods.py", "1–2 days", "Eng", "runs/calibration/best_params.json",
         "Measurable F1 lift ≥5% vs baseline", "Not started"),
        ("2", "P0", "Calibration", "Promote winning defaults",
         "Apply calibrated params as new defaults only where Delhi F1 improves and "
         "car-FP gate unchanged.",
         "Grid-search", "0.5 day", "Eng", "detection_config.py / constants",
         "Delhi mean F1 up; parked_cars F1=1.0", "Not started"),
        ("3", "P1", "Transfer learning", "Expand labeled set for training",
         "Grow Delhi labels to 100–300 pairs (semi-automated pre-label + human fix). "
         "Split train/val/test (70/15/15).",
         "Initial 20–40 pairs", "1–2 weeks", "Domain", "data/delhi_cd/train|val|test",
         "≥100 train pairs with masks", "Not started"),
        ("3", "P1", "Transfer learning", "Fine-tuning script",
         "scripts/finetune_adaptformer.py: load HF checkpoint, train on Delhi tiles, "
         "early-stop on val F1, export best weights.",
         "Expanded labels", "2–3 days", "Eng", "scripts/finetune_adaptformer.py",
         "Reproducible train run from CLI", "Not started"),
        ("3", "P1", "Transfer learning", "GPU training run",
         "Rent cloud GPU (A10/T4); train 20–50 epochs on 256px crops; log metrics.",
         "Fine-tuning script", "4–8 hours GPU", "Eng", "models/adaptformer_delhi/",
         "Val F1 beats pretrained on Delhi", "Not started"),
        ("3", "P1", "Transfer learning", "Integrate Delhi weights",
         "model_inference.py: load local fine-tuned weights when present "
         "(ADAPTFORMER_WEIGHTS env); fallback to HF LEVIR model.",
         "GPU run", "1 day", "Eng", "app/model_inference.py",
         "App uses Delhi model by default locally", "Not started"),
        ("4", "P2", "Optional DL", "BIT_CD ensemble A/B",
         "Only if fine-tuned AdaptFormer still misses specific change types; "
         "benchmark ensemble on Delhi eval before enabling.",
         "Phase 3 complete", "2 days", "Eng", "DETECTION_ENSEMBLE gate",
         "Delhi F1 gain ≥2% vs fine-tuned alone", "Deferred"),
        ("4", "P2", "Optional classical", "Extended KPCAMNet",
         "Multi-channel stacked KPCA, more components, domain-specific patch sizes. "
         "For unsupervised / no-GPU fallback only.",
         "Phase 2 complete", "3–5 days", "Eng", "app/cd_models/kpca_features.py",
         "Feature-Based Delhi F1 improves; AI path unaffected", "Deferred"),
        ("5", "P3", "Not planned", "Siamese U-Net from scratch",
         "Requires 1000+ labeled pairs; unlikely to beat fine-tuned AdaptFormer.",
         "—", "—", "—", "—", "Skip unless data scale changes", "Skipped"),
        ("6", "P1", "Production", "Re-run full validation gates",
         "Delhi eval + synthetic + car-FP + LEVIR regression after each promoted change.",
         "Phases 2–3", "0.5 day", "Eng", "runs/final_validation/",
         "All gates pass; metrics in README/docs", "Not started"),
        ("6", "P1", "Production", "Update docs & .env.example",
         "Document Delhi eval workflow, fine-tuned weights path, calibration commands.",
         "Validation", "0.5 day", "Eng", "DEV_SETUP.md, .env.example",
         "Team can reproduce benchmark end-to-end", "Not started"),
    ]
    write_table(ws, headers, rows, start_row=1, priority_col=2)
    autosize(ws)


def build_timeline(ws):
    ws.title = "Timeline"
    headers = ["Week", "Focus", "Key outputs", "Gate / decision point"]
    rows = [
        ("Week 1", "Delhi eval set + baseline",
         "20–40 labeled pairs; baseline metrics; compare_methods.py started",
         "Go/no-go: enough pairs to measure (≥20)"),
        ("Week 2", "Calibration",
         "Grid-search complete; promoted defaults; compare_methods leaderboard",
         "Go/no-go: ≥5% F1 lift or proceed to fine-tune anyway"),
        ("Week 2–4", "Label expansion + fine-tune",
         "100–300 pairs; GPU training; Delhi AdaptFormer weights",
         "Go/no-go: val F1 beats pretrained"),
        ("Week 4–5", "Integration + validation",
         "App loads Delhi weights; final gates; documentation",
         "Release calibrated + fine-tuned pipeline"),
        ("Week 5+ (optional)", "KPCAMNet / BIT_CD",
         "Only if fine-tuned model still insufficient",
         "Each optional item needs Delhi F1 proof"),
    ]
    write_table(ws, headers, rows)
    autosize(ws)


def build_risks(ws):
    ws.title = "Risks & Dependencies"
    headers = ["Risk", "Impact", "Mitigation", "Owner"]
    rows = [
        ("Insufficient labeled Delhi data", "Fine-tune underperforms",
         "Start with 20–40 for calibration; grow to 100+ before GPU spend",
         "Domain team"),
        ("Label inconsistency (rough polygons)", "Noisy F1, wrong tuning",
         "Labeling guide; double-review 10% of masks; focus on building/road changes",
         "Domain team"),
        ("CPU-only inference too slow", "Poor UX on large GeoTIFFs",
         "Keep 4096 cap; optional DETECTION_TTA=off; queue jobs with progress",
         "Eng"),
        ("GPU cost / access", "Blocks fine-tuning",
         "Single 4–8h cloud session; small tile dataset sufficient",
         "Eng"),
        ("Over-tuning on small eval set", "Overfits calibration",
         "Hold out 20% test pairs never used in grid-search",
         "Eng"),
        ("Car false positives return", "User trust loss",
         "Mandatory parked_cars synthetic gate before any default change",
         "Eng"),
    ]
    write_table(ws, headers, rows)
    autosize(ws)


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_summary(wb.active)
    build_comparison(wb.create_sheet())
    build_plan(wb.create_sheet())
    build_timeline(wb.create_sheet())
    build_risks(wb.create_sheet())
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
